#!/usr/bin/env python3
"""
Import historical pre-2024 student data from the SSA Dropbox XLS archive
and emit a single JSON that the dashboard serves via /api/historical-students.

Inputs:
  - SSA/Storico corsi/Lista Sommelier/Lista Sake Sommelier.xlsx  (passed, 2016-2021)
  - SSA/Storico corsi/LISTA BOCCIATI/BOCCIATI.xlsx               (failed)
  - SSA/Storico corsi/liste corsi eventbrite/Lista tutti corsi eventbrite certificato.xlsx
                                                                  (Eventbrite enrollments)

Outputs:
  - data/historical-students.json

Dedupe strategy:
  Entries are keyed by (email_lower, course_date_year). Passed/failed beats
  "enrolled" — so if someone has an Eventbrite enrollment AND a passed/failed
  record, the result entry wins.

Run from repo root:
    python3 scripts/import-historical.py
"""
from __future__ import annotations
import json
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install --user openpyxl")
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
DROPBOX_ROOT = Path(os.path.expanduser(
    "~/The WishList Dropbox/lorenzo ferraboschi/SSA/Storico corsi"
))
OUTPUT = REPO_ROOT / "data" / "historical-students.json"


def excel_serial_to_date(value):
    """Excel stores dates as serial numbers sometimes. Convert to date()."""
    if isinstance(value, (datetime, date)):
        return value if isinstance(value, date) else value.date()
    if isinstance(value, (int, float)):
        # Excel epoch is 1899-12-30 (accounts for the 1900 leap-year bug)
        try:
            return date(1899, 12, 30) + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        s = value.strip()
        # Try ISO
        m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
        if m:
            try:
                return date(int(m[1]), int(m[2]), int(m[3]))
            except ValueError:
                return None
        # Italian DD/MM/YYYY or DD/MM/YY
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
        if m:
            d, mo, y = int(m[1]), int(m[2]), int(m[3])
            if y < 100:
                y = 2000 + y if y < 50 else 1900 + y
            try:
                return date(y, mo, d)
            except ValueError:
                return None
    return None


def clean_email(raw):
    if not raw:
        return ''
    e = str(raw).strip().lower()
    # Defensive: drop obvious garbage
    if '@' not in e or ' ' in e:
        return ''
    return e


def clean_name(raw):
    if not raw:
        return ''
    return ' '.join(str(raw).strip().split())


def clean_phone(raw):
    if not raw:
        return ''
    p = re.sub(r'\s+', '', str(raw))
    return p if len(p) >= 6 else ''


def title_case_if_shouty(name):
    """Input like 'MARINA MARMIROLI' -> 'Marina Marmiroli'. Leaves mixed-case alone."""
    if not name:
        return name
    if name.isupper() or name.islower():
        return ' '.join(part.capitalize() for part in name.split())
    return name


def guess_city(value):
    if not value:
        return ''
    v = str(value).strip()
    # Keep as-is; the frontend can normalize display
    return v


def load_passed(path: Path):
    """Lista Sake Sommelier.xlsx -> list of passed students (2016-2021)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header row
            continue
        if not row or not any(v not in (None, '') for v in row[:4]):
            continue
        cognome = clean_name(row[0]) if len(row) > 0 else ''
        nome = clean_name(row[1]) if len(row) > 1 else ''
        email = clean_email(row[2]) if len(row) > 2 else ''
        d = excel_serial_to_date(row[3]) if len(row) > 3 else None
        if not (nome or cognome) and not email:
            continue
        out.append({
            'firstName': title_case_if_shouty(nome),
            'lastName': title_case_if_shouty(cognome),
            'email': email,
            'phone': '',
            'location': '',
            'courseDate': d.isoformat() if d else None,
            'result': 'passed',
            'source': 'lista-sommelier-xlsx',
        })
    wb.close()
    return out


def load_failed(path: Path):
    """BOCCIATI.xlsx -> list of failed students."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header row
            continue
        if not row or not any(v not in (None, '') for v in row[:12]):
            continue
        # Cols per R0 in earlier inspection:
        # 0 Course, 1 Franchise, 2 Educator, 3 Course Date, 4 Exam Date,
        # 5 Venue, 6 Gender, 7 First Name, 8 Last Name, ..., 11 E-Mail
        course_date = excel_serial_to_date(row[3]) if len(row) > 3 else None
        exam_date = excel_serial_to_date(row[4]) if len(row) > 4 else None
        venue = guess_city(row[5]) if len(row) > 5 else ''
        first = clean_name(row[7]) if len(row) > 7 else ''
        last = clean_name(row[8]) if len(row) > 8 else ''
        email = clean_email(row[11]) if len(row) > 11 else ''
        if not (first or last) and not email:
            continue
        out.append({
            'firstName': title_case_if_shouty(first),
            'lastName': title_case_if_shouty(last),
            'email': email,
            'phone': '',
            'location': venue,
            'courseDate': course_date.isoformat() if course_date else None,
            'examDate': exam_date.isoformat() if exam_date else None,
            'result': 'failed',
            'source': 'bocciati-xlsx',
        })
    wb.close()
    return out


def load_eventbrite(path: Path):
    """Lista tutti corsi eventbrite certificato.xlsx -> enrolled students."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not any(v not in (None, '') for v in row[:5]):
            continue
        # Header: DATA CORSO, NOME, COGNOME, MAIL, TEL, NASCITA, INDIRIZZO, ..., (city somewhere)
        course_date = excel_serial_to_date(row[0]) if len(row) > 0 else None
        nome = clean_name(row[1]) if len(row) > 1 else ''
        cognome = clean_name(row[2]) if len(row) > 2 else ''
        email = clean_email(row[3]) if len(row) > 3 else ''
        phone = clean_phone(row[4]) if len(row) > 4 else ''
        city = ''
        if len(row) > 8 and row[8]:
            city = guess_city(row[8])
        if not (nome or cognome) and not email:
            continue
        out.append({
            'firstName': title_case_if_shouty(nome),
            'lastName': title_case_if_shouty(cognome),
            'email': email,
            'phone': phone,
            'location': city,
            'courseDate': course_date.isoformat() if course_date else None,
            'result': 'enrolled',  # Eventbrite list gives enrollment, not outcome
            'source': 'eventbrite-xlsx',
        })
    wb.close()
    return out


def merge_and_dedupe(*lists):
    """Merge multiple lists; dedupe on (email_lower, courseDate_year).

    Priority: passed > failed > enrolled.
    """
    priority = {'passed': 3, 'failed': 2, 'enrolled': 1}
    key_to_entry = {}

    for source_list in lists:
        for e in source_list:
            email = e.get('email') or ''
            date_str = e.get('courseDate') or ''
            year = date_str[:4] if date_str else 'unknown'
            # Key by email+year, or by normalized name+year if no email
            if email:
                key = (email, year)
            else:
                name_key = f"{e.get('firstName','').lower()}|{e.get('lastName','').lower()}"
                key = (name_key, year)
            existing = key_to_entry.get(key)
            if not existing or priority.get(e.get('result'), 0) > priority.get(existing.get('result'), 0):
                key_to_entry[key] = e
            elif priority.get(e.get('result'), 0) == priority.get(existing.get('result'), 0):
                # Same level — keep the one with more data (phone/city)
                if (e.get('phone') and not existing.get('phone')) or \
                   (e.get('location') and not existing.get('location')):
                    key_to_entry[key] = e
    return list(key_to_entry.values())


def main():
    if not DROPBOX_ROOT.exists():
        print(f"ERROR: Dropbox path not found: {DROPBOX_ROOT}")
        sys.exit(1)

    passed_path = DROPBOX_ROOT / "Lista Sommelier" / "Lista Sake Sommelier.xlsx"
    failed_path = DROPBOX_ROOT / "LISTA BOCCIATI" / "BOCCIATI.xlsx"
    event_path = DROPBOX_ROOT / "liste corsi eventbrite" / "Lista tutti corsi eventbrite certificato.xlsx"

    print(f"Reading: {passed_path.name} ...")
    passed = load_passed(passed_path)
    print(f"  {len(passed)} passed records")

    print(f"Reading: {failed_path.name} ...")
    failed = load_failed(failed_path)
    print(f"  {len(failed)} failed records")

    print(f"Reading: {event_path.name} ...")
    events = load_eventbrite(event_path)
    print(f"  {len(events)} eventbrite enrollments")

    merged = merge_and_dedupe(passed, failed, events)
    print(f"\nAfter dedupe: {len(merged)} unique entries")

    # Sort by courseDate desc, then name
    merged.sort(key=lambda e: (
        e.get('courseDate') or '',
        e.get('lastName') or '',
        e.get('firstName') or ''
    ), reverse=False)

    # Stats
    by_result = {}
    by_year = {}
    for e in merged:
        by_result[e['result']] = by_result.get(e['result'], 0) + 1
        y = (e.get('courseDate') or 'unknown')[:4]
        by_year[y] = by_year.get(y, 0) + 1

    print(f"By result: {by_result}")
    print(f"By year:   {dict(sorted(by_year.items()))}")

    output = {
        'generatedAt': datetime.utcnow().isoformat() + 'Z',
        'count': len(merged),
        'byResult': by_result,
        'byYear': by_year,
        'students': merged,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nWrote {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == '__main__':
    main()
