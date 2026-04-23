#!/usr/bin/env python3
"""
Bulk-import exam outcomes from a Socrative XLS export into the SSA dashboard.

Each XLS lists every student who took the exam for a given course. Per the
business rule ("chi appare nell'XLS è promosso"), every row with a valid email
gets an 'passed' override on the target Shopify course. Low-score attempts can
be skipped with --min-score to avoid ingesting test/incomplete rows.

Typical use:
    python3 scripts/import-xls-outcomes.py \
        /path/to/Class_..._NIHONSHU-2603-MILANO.xlsx \
        --handle corso-di-sake-sommelier-certificato-marzo-2026-milano \
        --url https://corsi.sakesommelierassociation.it \
        --user admin --password $SSA_PASS \
        --min-score 50

Dry run (no API calls) to preview the import:
    python3 scripts/import-xls-outcomes.py ... --dry-run

Dependencies: openpyxl, requests
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    print('Missing dependency: pip install openpyxl', file=sys.stderr)
    sys.exit(2)

try:
    import requests  # type: ignore
except ImportError:
    print('Missing dependency: pip install requests', file=sys.stderr)
    sys.exit(2)


# Column layout of Socrative exam exports (observed on 2026-04 files):
#   row 0 col 0  -> exam code (e.g. "NIHONSHU-2603-MILANO")
#   row 1 col 0  -> exam datetime (human readable)
#   row 5        -> question headers
#   row 6        -> points-per-question definitions
#   rows 7..n    -> student rows
#   row n        -> "Punteggio di classe" (aggregate, skip)
# Student row fields:
#   c0 name, c1 student id ("-"), c2 score %, c3 points, c4 typed name,
#   c5 date of birth, c6 nationality, c7 phone, c8 email, c9 address, ...
SCORE_COL = 2
TYPED_NAME_COL = 4
EMAIL_COL = 8


@dataclass
class StudentRow:
    name: str
    email: str
    score: float

    @property
    def email_key(self) -> str:
        return (self.email or '').lower().strip()


def parse_xls(path: Path) -> tuple[str, list[StudentRow]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0] or not rows[0][0]:
        raise SystemExit(f'Unexpected XLS layout (empty header row): {path}')
    course_code = str(rows[0][0]).strip()
    students: list[StudentRow] = []
    # Students rows start at index 7 (after header + points definition).
    for i, row in enumerate(rows[7:], start=7):
        if not row:
            continue
        first = (row[0] or '').__str__().strip()
        if not first or first.lower().startswith('punteggio di classe'):
            break
        score_raw = row[SCORE_COL] if len(row) > SCORE_COL else None
        try:
            score = float(score_raw) if score_raw not in (None, '') else 0.0
        except (TypeError, ValueError):
            score = 0.0
        typed_name = (row[TYPED_NAME_COL] if len(row) > TYPED_NAME_COL else '') or ''
        email = (row[EMAIL_COL] if len(row) > EMAIL_COL else '') or ''
        name = (typed_name or first).strip()
        email_s = str(email).strip()
        students.append(StudentRow(name=name, email=email_s, score=score))
    return course_code, students


def dedupe_keep_best(students: list[StudentRow]) -> list[StudentRow]:
    """Multiple attempts collapse into one row per email, keeping the highest score."""
    best: dict[str, StudentRow] = {}
    for s in students:
        if not s.email_key:
            continue
        cur = best.get(s.email_key)
        if not cur or s.score > cur.score:
            best[s.email_key] = s
    return list(best.values())


def login(url: str, user: str, password: str) -> requests.Session:
    sess = requests.Session()
    resp = sess.post(
        f'{url.rstrip("/")}/auth/login',
        json={'username': user, 'password': password},
        timeout=15,
    )
    if resp.status_code != 200:
        raise SystemExit(f'Login failed: {resp.status_code} {resp.text}')
    return sess


def push_outcome(sess: requests.Session, url: str, handle: str, email: str, outcome: str) -> tuple[bool, str]:
    endpoint = f'{url.rstrip("/")}/api/costs/{handle}/outcome/{email}'
    try:
        resp = sess.post(endpoint, json={'outcome': outcome}, timeout=15)
        if resp.status_code != 200:
            return False, f'{resp.status_code} {resp.text[:120]}'
        return True, 'ok'
    except requests.RequestException as e:
        return False, str(e)


def main() -> int:
    ap = argparse.ArgumentParser(description='Import Socrative XLS outcomes into SSA dashboard')
    ap.add_argument('xls', type=Path, help='Path to the Socrative XLS file')
    ap.add_argument('--handle', required=True, help='Shopify course handle to assign overrides to')
    ap.add_argument('--url', default='https://corsi.sakesommelierassociation.it', help='Dashboard base URL')
    ap.add_argument('--user', default='admin', help='Dashboard username')
    ap.add_argument('--password', help='Dashboard password (or $SSA_PASS)')
    ap.add_argument('--min-score', type=float, default=50.0,
                    help='Skip rows with score below this threshold (default: 50). Use 0 to import everyone.')
    ap.add_argument('--outcome', default='passed', choices=['passed', 'failed'],
                    help='Outcome to set on each imported student (default: passed)')
    ap.add_argument('--dry-run', action='store_true', help='Preview the import without calling the API')
    args = ap.parse_args()

    if not args.xls.exists():
        raise SystemExit(f'File not found: {args.xls}')

    code, rows = parse_xls(args.xls)
    keep = [s for s in dedupe_keep_best(rows) if s.score >= args.min_score]
    dropped = [s for s in dedupe_keep_best(rows) if s.score < args.min_score]

    print(f'File:          {args.xls.name}')
    print(f'Exam code:     {code}')
    print(f'Target handle: {args.handle}')
    print(f'Students (valid email, score >= {args.min_score}): {len(keep)}')
    if dropped:
        print(f'  (skipping {len(dropped)} rows below threshold):')
        for s in dropped:
            print(f'    - {s.name!r:30s} {s.email_key} score={s.score}')

    print('Students to upsert:')
    for s in keep:
        print(f'  {s.name!r:30s} {s.email_key:40s} score={s.score}')

    if args.dry_run:
        print('\n[DRY RUN] Nothing posted.')
        return 0

    import os
    password = args.password or os.environ.get('SSA_PASS')
    if not password:
        raise SystemExit('Password required: pass --password or set $SSA_PASS')

    sess = login(args.url, args.user, password)
    ok = 0
    fail: list[tuple[StudentRow, str]] = []
    for s in keep:
        success, msg = push_outcome(sess, args.url, args.handle, s.email_key, args.outcome)
        if success:
            ok += 1
            print(f'  [OK]   {s.email_key} -> {args.outcome}')
        else:
            fail.append((s, msg))
            print(f'  [FAIL] {s.email_key}: {msg}')

    print(f'\nDone: {ok} saved, {len(fail)} failed.')
    return 0 if not fail else 1


if __name__ == '__main__':
    sys.exit(main())
